"""
Parser de archivos Excel de oferta de granjas.
Lee el formato de la pestaña 'OFERTA JUEV' y lo convierte en LoteOferta.
Soporta tanto .xlsx (openpyxl) como .xls (xlrd).
"""
import openpyxl
import xlrd
from datetime import date, datetime
from typing import List, Optional, BinaryIO
from io import BytesIO

from .calculo import LoteOferta


# Mapeo de columnas esperadas en la oferta (pestaña OFERTA JUEV / hoja de oferta)
COLUMNAS_OFERTA = {
    "fecha_peso": 0,       # A - Fecha de Peso
    "granja": 1,           # B - GRANJA
    "galpon": 2,           # C - Galpon
    "nucleo": 3,           # D - Nucleo
    "cantidad": 4,         # E - cantidad
    "sexo": 5,             # F - sexo
    "edad_proyectada": 6,  # G - Edad Proyectada
    "peso_muestreo_proy": 7,  # H - Peso Muestreo Proy
    "ganancia_diaria": 8,  # I - Ganancia Diaria
    "dias_proyectados": 9, # J - Dias proyectados
    "edad_real": 10,       # K - EDAD REAL
    "peso_muestreo_real": 11,  # L - Peso Muestreo REAL
    "fecha_ingreso": 12,   # M - FECHA DE INGRESO
}

FILA_INICIO_DATOS = 4  # Fila donde empiezan los datos (1-indexed, fila 4 en Excel)


def _parse_date(val) -> Optional[date]:
    """Convierte un valor de celda a date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _parse_int(val) -> int:
    """Convierte un valor de celda a int."""
    if val is None:
        return 0
    try:
        if isinstance(val, (int, float)):
            return int(val)
        s = str(val).strip()
        # Si tiene coma como separador decimal (ej: "4,5")
        s = s.replace(",", ".")
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def _parse_float(val) -> float:
    """Convierte un valor de celda a float."""
    if val is None:
        return 0.0
    try:
        s = str(val).replace(",", ".")
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _parse_sexo(val) -> str:
    """Convierte un valor de celda a código de sexo."""
    if val is None:
        return ""
    s = str(val).strip().upper()
    if s in ("M", "MACHO"):
        return "M"
    elif s in ("H", "HEMBRA"):
        return "H"
    elif s in ("MIX", "MIXTO"):
        return "MIX"
    return ""


def leer_oferta_excel(
    file_content: bytes,
    sheet_name: Optional[str] = None
) -> List[LoteOferta]:
    """
    Lee un archivo Excel de oferta y devuelve la lista de LoteOferta.
    
    Args:
        file_content: contenido binario del archivo Excel
        sheet_name: nombre de la pestaña (si None, usa la primera o busca 'OFERTA')
    
    Returns:
        Lista de LoteOferta parseados
    """
    # Detectar formato: intentar primero con openpyxl (.xlsx)
    # Si falla, intentar con xlrd (.xls)
    is_xls = False
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    except Exception:
        # Probablemente es un .xls (formato binario antiguo)
        try:
            wb = xlrd.open_workbook(file_contents=file_content)
            is_xls = True
        except Exception as e:
            raise ValueError(f"No se pudo abrir el archivo Excel: {e}")

    if is_xls:
        return _leer_oferta_xlrd(wb, sheet_name)
    else:
        return _leer_oferta_openpyxl(wb, sheet_name)


def _leer_oferta_openpyxl(wb, sheet_name: Optional[str]) -> List[LoteOferta]:
    """Lee oferta desde un workbook de openpyxl (.xlsx)."""
    # Buscar la pestaña adecuada
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Buscar una pestaña que contenga "OFERTA"
        oferta_sheet = None
        for name in wb.sheetnames:
            if "OFERTA" in name.upper():
                oferta_sheet = name
                break
        if oferta_sheet:
            ws = wb[oferta_sheet]
        else:
            ws = wb.active

    lotes: List[LoteOferta] = []

    for row_idx in range(FILA_INICIO_DATOS, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c + 1).value for c in range(14)]
        lote = _parse_row_to_lote(row)
        if lote:
            lotes.append(lote)

    wb.close()
    return lotes


def _xlrd_cell_value(sheet, row, col):
    """Obtiene el valor de una celda xlrd, convirtiendo fechas."""
    if col >= sheet.ncols:
        return None
    cell = sheet.cell(row, col)
    # xlrd cell types: 0=empty, 1=text, 2=number, 3=date, 4=boolean, 5=error, 6=blank
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            dt_tuple = xlrd.xldate_as_tuple(cell.value, sheet.book.datemode)
            return datetime(*dt_tuple)
        except Exception:
            return cell.value
    if cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
        return None
    return cell.value


def _leer_oferta_xlrd(wb, sheet_name: Optional[str]) -> List[LoteOferta]:
    """Lee oferta desde un workbook de xlrd (.xls)."""
    # Buscar la pestaña adecuada
    if sheet_name and sheet_name in wb.sheet_names():
        ws = wb.sheet_by_name(sheet_name)
    else:
        oferta_sheet = None
        for name in wb.sheet_names():
            if "OFERTA" in name.upper():
                oferta_sheet = name
                break
        if oferta_sheet:
            ws = wb.sheet_by_name(oferta_sheet)
        else:
            ws = wb.sheet_by_index(0)

    lotes: List[LoteOferta] = []

    # xlrd usa índices 0-based; FILA_INICIO_DATOS es 4 (1-indexed) → fila 3 en 0-indexed
    for row_idx in range(FILA_INICIO_DATOS - 1, ws.nrows):
        row = [_xlrd_cell_value(ws, row_idx, c) for c in range(14)]
        lote = _parse_row_to_lote(row)
        if lote:
            lotes.append(lote)

    return lotes


def _parse_row_to_lote(row: list) -> Optional[LoteOferta]:
    """Convierte una fila de datos a LoteOferta. Retorna None si la fila es inválida."""
    granja = row[COLUMNAS_OFERTA["granja"]]
    if not granja or str(granja).strip() == "":
        return None

    fecha_peso = _parse_date(row[COLUMNAS_OFERTA["fecha_peso"]])
    fecha_ingreso = _parse_date(row[COLUMNAS_OFERTA["fecha_ingreso"]])

    if not fecha_peso:
        return None

    cantidad_raw = row[COLUMNAS_OFERTA["cantidad"]]
    cantidad = _parse_int(cantidad_raw)
    # Si la cantidad tiene punto de miles (ej: 15.685 → 15685)
    if isinstance(cantidad_raw, float) and cantidad_raw > 100:
        cantidad = int(cantidad_raw)
    elif isinstance(cantidad_raw, (int, float)):
        cantidad = int(cantidad_raw)

    return LoteOferta(
        fecha_peso=fecha_peso,
        granja=str(granja).strip(),
        galpon=_parse_int(row[COLUMNAS_OFERTA["galpon"]]),
        nucleo=_parse_int(row[COLUMNAS_OFERTA["nucleo"]]),
        cantidad=cantidad,
        sexo=_parse_sexo(row[COLUMNAS_OFERTA["sexo"]]),
        edad_proyectada=_parse_int(row[COLUMNAS_OFERTA["edad_proyectada"]]),
        peso_muestreo_proy=_parse_float(row[COLUMNAS_OFERTA["peso_muestreo_proy"]]),
        ganancia_diaria=_parse_float(row[COLUMNAS_OFERTA["ganancia_diaria"]]),
        dias_proyectados=_parse_int(row[COLUMNAS_OFERTA["dias_proyectados"]]),
        edad_real=_parse_int(row[COLUMNAS_OFERTA["edad_real"]]),
        peso_muestreo_real=_parse_float(row[COLUMNAS_OFERTA["peso_muestreo_real"]]),
        fecha_ingreso=fecha_ingreso or fecha_peso,
    )


def leer_proyeccion_excel(
    file_content: bytes,
    sheet_name: str = "PROYEC1"
) -> dict:
    """
    Lee la pestaña PROYEC1 del Excel para extraer la configuración existente.
    Retorna los parámetros y la estructura de la proyección.
    Soporta tanto .xlsx (openpyxl) como .xls (xlrd).
    """
    is_xls = False
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    except Exception:
        try:
            wb = xlrd.open_workbook(file_contents=file_content)
            is_xls = True
        except Exception as e:
            raise ValueError(f"No se pudo abrir el archivo Excel: {e}")

    if is_xls:
        if sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
        else:
            ws = wb.sheet_by_index(0)
        resultado = {
            "fecha_base": _parse_date(_xlrd_cell_value(ws, 3, 3)),   # D4 (0-indexed: row=3, col=3)
            "ganancia_macho": _parse_float(_xlrd_cell_value(ws, 4, 13)),   # N5
            "ganancia_hembra": _parse_float(_xlrd_cell_value(ws, 4, 15)),  # P5
        }
    else:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        resultado = {
            "fecha_base": _parse_date(ws.cell(row=4, column=4).value),
            "ganancia_macho": _parse_float(ws.cell(row=5, column=14).value),  # N5
            "ganancia_hembra": _parse_float(ws.cell(row=5, column=16).value),  # P5
        }
        wb.close()

    return resultado
