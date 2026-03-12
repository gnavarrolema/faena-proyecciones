"""
Parser del archivo Excel "13.Datos Produccion por Semana".

Lee las columnas:
- B/C: Semana (DESDE/HASTA) — la columna A está vacía
- I: Pollitos Cargados en Granjas Propias

Y genera simulaciones de mortalidad a diferentes tasas.
"""
import openpyxl
import xlrd
from datetime import date, datetime, timedelta
from typing import List, Optional
from io import BytesIO
from pydantic import BaseModel


# ─── Modelos ────────────────────────────────────────────────────────────────────

class SemanaProduccion(BaseModel):
    """Datos de producción de una semana."""
    fecha_desde: date
    fecha_hasta: date
    pollitos_cargados: int  # columna I: Pollitos Cargados en Granjas Propias


class SimulacionMortalidadFila(BaseModel):
    """Resultado de la simulación de mortalidad para una semana y una tasa."""
    tasa_mortalidad: float         # ej: 0.065 (6.5%)
    pollitos_disponibles: int      # pollitos_cargados * (1 - tasa)


class SemanaSimulacion(BaseModel):
    """Simulación completa para una semana con múltiples tasas de mortalidad."""
    fecha_desde: date
    fecha_hasta: date
    pollitos_cargados: int
    fecha_faena_estimada: date     # fecha_desde + 42 días
    simulaciones: List[SimulacionMortalidadFila]


# ─── Constantes ─────────────────────────────────────────────────────────────────

DIAS_HASTA_FAENA = 42

TASAS_MORTALIDAD_DEFAULT = [0.045, 0.05, 0.055, 0.06, 0.065]

# Mapeo de columnas del Excel de producción
# Columna A está vacía; los datos empiezan en B
COL_DESDE = 1       # B - DESDE (0-indexed)
COL_HASTA = 2       # C - HASTA (0-indexed, se auto-detecta si difiere)
COL_POLLITOS = 8    # I - Pollitos Cargados en Granjas Propias (0-indexed = 8)

FILA_INICIO_DATOS = 6  # Fila donde empiezan los datos (1-indexed, fila 6)


# ─── Helpers ────────────────────────────────────────────────────────────────────

def _parse_date(val) -> Optional[date]:
    """Convierte un valor de celda a date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _parse_int(val) -> int:
    """Convierte un valor de celda a int."""
    if val is None:
        return 0
    try:
        if isinstance(val, (int, float)):
            return int(val)
        s = str(val).strip().replace(",", ".").replace(".", "")
        return int(float(s)) if s else 0
    except (ValueError, TypeError):
        return 0


# ─── Parser ─────────────────────────────────────────────────────────────────────

def leer_produccion_excel(
    file_content: bytes,
    sheet_name: Optional[str] = None,
) -> List[SemanaProduccion]:
    """
    Lee el Excel de producción semanal.

    Busca la pestaña que contenga "DATOS" o "PRODUCCION" o usa la primera.
    Lee columnas B (DESDE), C (HASTA) y I (Pollitos Cargados).
    """
    is_xls = False
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    except Exception:
        try:
            wb = xlrd.open_workbook(file_contents=file_content)
            is_xls = True
        except Exception as e:
            raise ValueError(f"No se pudo abrir el archivo Excel: {e}")

    if is_xls:
        return _leer_produccion_xlrd(wb, sheet_name)
    else:
        return _leer_produccion_openpyxl(wb, sheet_name)


def _buscar_hoja_produccion_openpyxl(wb, sheet_name: Optional[str]):
    """Busca la hoja correcta en un workbook openpyxl."""
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]

    for name in wb.sheetnames:
        upper = name.upper()
        if "DATOS" in upper or "PRODUCCION" in upper or "PRODUCCIÓN" in upper:
            return wb[name]

    return wb.active


def _buscar_hoja_produccion_xlrd(wb, sheet_name: Optional[str]):
    """Busca la hoja correcta en un workbook xlrd."""
    if sheet_name and sheet_name in wb.sheet_names():
        return wb.sheet_by_name(sheet_name)

    for name in wb.sheet_names():
        upper = name.upper()
        if "DATOS" in upper or "PRODUCCION" in upper or "PRODUCCIÓN" in upper:
            return wb.sheet_by_name(name)

    return wb.sheet_by_index(0)


def _detectar_columna_hasta(ws_openpyxl=None, ws_xlrd=None, fila_header: int = 4) -> int:
    """
    Detecta si la columna HASTA es B(1) o C(2), buscando en el header.
    Por defecto retorna 1 (columna B).
    """
    if ws_openpyxl:
        for col in range(1, 4):  # Columnas A-C (1-indexed)
            val = ws_openpyxl.cell(row=fila_header, column=col + 1).value
            if val and "HASTA" in str(val).upper():
                return col  # 0-indexed
    elif ws_xlrd:
        for col in range(3):  # Columnas A-C (0-indexed)
            if col < ws_xlrd.ncols:
                val = ws_xlrd.cell_value(fila_header - 1, col)
                if val and "HASTA" in str(val).upper():
                    return col
    return 1  # Default: columna B (0-indexed)


def _leer_produccion_openpyxl(wb, sheet_name: Optional[str]) -> List[SemanaProduccion]:
    """Lee producción desde workbook openpyxl (.xlsx)."""
    ws = _buscar_hoja_produccion_openpyxl(wb, sheet_name)

    col_hasta = _detectar_columna_hasta(ws_openpyxl=ws)
    semanas: List[SemanaProduccion] = []

    for row_idx in range(FILA_INICIO_DATOS, ws.max_row + 1):
        fecha_desde = _parse_date(ws.cell(row=row_idx, column=COL_DESDE + 1).value)
        fecha_hasta = _parse_date(ws.cell(row=row_idx, column=col_hasta + 1).value)
        pollitos = _parse_int(ws.cell(row=row_idx, column=COL_POLLITOS + 1).value)

        if not fecha_desde or not fecha_hasta:
            continue
        if pollitos <= 0:
            continue

        semanas.append(SemanaProduccion(
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            pollitos_cargados=pollitos,
        ))

    wb.close()
    return semanas


def _xlrd_cell_value(sheet, row, col):
    """Obtiene el valor de una celda xlrd, convirtiendo fechas."""
    if col >= sheet.ncols:
        return None
    cell = sheet.cell(row, col)
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            dt_tuple = xlrd.xldate_as_tuple(cell.value, sheet.book.datemode)
            return datetime(*dt_tuple)
        except Exception:
            return cell.value
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    return cell.value


def _leer_produccion_xlrd(wb, sheet_name: Optional[str]) -> List[SemanaProduccion]:
    """Lee producción desde workbook xlrd (.xls)."""
    ws = _buscar_hoja_produccion_xlrd(wb, sheet_name)

    col_hasta = _detectar_columna_hasta(ws_xlrd=ws)
    semanas: List[SemanaProduccion] = []

    for row_idx in range(FILA_INICIO_DATOS - 1, ws.nrows):
        fecha_desde = _parse_date(_xlrd_cell_value(ws, row_idx, COL_DESDE))
        fecha_hasta = _parse_date(_xlrd_cell_value(ws, row_idx, col_hasta))
        pollitos = _parse_int(_xlrd_cell_value(ws, row_idx, COL_POLLITOS))

        if not fecha_desde or not fecha_hasta:
            continue
        if pollitos <= 0:
            continue

        semanas.append(SemanaProduccion(
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            pollitos_cargados=pollitos,
        ))

    return semanas


# ─── Simulación de mortalidad ───────────────────────────────────────────────────

def simular_mortalidad(
    semanas: List[SemanaProduccion],
    tasas: Optional[List[float]] = None,
) -> List[SemanaSimulacion]:
    """
    Para cada semana de producción, calcula cuántos pollitos estarán
    disponibles para faena aplicando cada tasa de mortalidad.

    La fecha de faena estimada es fecha_desde + 42 días.
    """
    if tasas is None:
        tasas = TASAS_MORTALIDAD_DEFAULT

    resultado: List[SemanaSimulacion] = []

    for sem in semanas:
        fecha_faena = sem.fecha_desde + timedelta(days=DIAS_HASTA_FAENA)

        simulaciones = []
        for tasa in tasas:
            disponibles = int(sem.pollitos_cargados * (1 - tasa))
            simulaciones.append(SimulacionMortalidadFila(
                tasa_mortalidad=tasa,
                pollitos_disponibles=disponibles,
            ))

        resultado.append(SemanaSimulacion(
            fecha_desde=sem.fecha_desde,
            fecha_hasta=sem.fecha_hasta,
            pollitos_cargados=sem.pollitos_cargados,
            fecha_faena_estimada=fecha_faena,
            simulaciones=simulaciones,
        ))

    return resultado
