"""Compara la proyección actual del backend contra la planilla PROYEC 3.

Uso rápido:
    python scripts/compare_manager_heuristics.py
    python scripts/compare_manager_heuristics.py --keys MANANTIALES:1:2:H MANANTIALES:2:2:H
    python scripts/compare_manager_heuristics.py --priorizar-peso-objetivo
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

import openpyxl

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from backend.calculo import Parametros, generar_proyeccion, normalizar_granja_clave
from backend.main import ProyeccionRequest, _resolver_calendario_planificacion
from backend.parser_excel import leer_oferta_excel


DEFAULT_OFFER = BASE_DIR / "Anexos" / "OFERTA DEL 16-4-26.xls"
DEFAULT_MANAGER = BASE_DIR / "Anexos" / "Proyeccion de faena 16-04-2026.xlsx"
DEFAULT_PARAMS = BASE_DIR / "local_storage" / "parametros.json"


@dataclass(frozen=True)
class LoteClave:
    granja: str
    galpon: int
    nucleo: int
    sexo: str

    def label(self) -> str:
        return f"{self.granja}:{self.galpon}:{self.nucleo}:{self.sexo}"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compara backend vs PROYEC 3 para el modo gerente.")
    parser.add_argument("--offer", type=Path, default=DEFAULT_OFFER, help="Archivo de oferta XLS/XLSX")
    parser.add_argument("--manager", type=Path, default=DEFAULT_MANAGER, help="Workbook con PROYEC 3")
    parser.add_argument("--sheet", default="PROYEC 3", help="Pestaña del workbook gerente")
    parser.add_argument("--params", type=Path, default=DEFAULT_PARAMS, help="JSON de parámetros persistidos")
    parser.add_argument("--fecha-inicio", default="2026-04-20", help="Fecha inicio semana YYYY-MM-DD")
    parser.add_argument("--dias-faena", type=int, default=5, help="Días de faena solicitados")
    parser.add_argument("--pollos-por-dia", type=int, default=35000, help="Objetivo diario solicitado")
    parser.add_argument("--keys", nargs="*", default=[], help="Claves GRANJA:GALPON:NUCLEO:SEXO a mostrar")
    parser.add_argument("--top", type=int, default=12, help="Cantidad de lotes a mostrar si no se pasan keys")
    parser.add_argument(
        "--priorizar-peso-objetivo",
        action="store_true",
        help="Activa el tie-break opcional por cercanía a peso objetivo entre lotes limpios.",
    )
    return parser.parse_args()


def parse_key(raw: str) -> LoteClave:
    granja, galpon, nucleo, sexo = raw.split(":", 3)
    return LoteClave(normalizar_granja_clave(granja), int(galpon), int(nucleo), sexo.upper())


def load_params(path: Path, priorizar_peso_objetivo: bool) -> Parametros:
    data = {}
    if path.exists():
        data = json.loads(path.read_text(encoding="utf-8"))
    data["planificacion_continua_gerente"] = True
    data.setdefault("planificacion_continua_dias_habiles", 16)
    data["planificacion_gerente_priorizar_peso_objetivo"] = priorizar_peso_objetivo
    return Parametros(**data)


def load_offer(path: Path):
    ofertas, _, _ = leer_oferta_excel(path.read_bytes())
    return ofertas


def make_key(granja: str, galpon: int, nucleo: int, sexo: str) -> LoteClave:
    return LoteClave(normalizar_granja_clave(granja), int(galpon), int(nucleo), (sexo or "").upper())


def load_manager_rows(path: Path, sheet_name: str) -> dict[LoteClave, list[dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    grouped: dict[LoteClave, list[dict]] = defaultdict(list)
    for row in range(8, ws.max_row + 1):
        granja = ws.cell(row, 1).value
        galpon = ws.cell(row, 2).value
        nucleo = ws.cell(row, 3).value
        cantidad = ws.cell(row, 4).value
        sexo = ws.cell(row, 5).value
        fecha = ws.cell(row, 8).value
        edad = ws.cell(row, 9).value
        peso = ws.cell(row, 12).value
        if not granja or galpon is None or nucleo is None or not sexo or not fecha:
            continue
        if cantidad in (None, 0):
            continue
        clave = make_key(str(granja), galpon, nucleo, str(sexo))
        grouped[clave].append(
            {
                "fecha": fecha.date() if hasattr(fecha, "date") else fecha,
                "cantidad": int(cantidad),
                "edad": int(edad) if edad is not None else None,
                "peso": float(peso) if peso is not None else None,
            }
        )
    for rows in grouped.values():
        rows.sort(key=lambda item: (item["fecha"], item["cantidad"]))
    return grouped


def build_projection(args: argparse.Namespace, params: Parametros):
    ofertas = load_offer(args.offer)
    req = ProyeccionRequest(
        fecha_inicio_semana=date.fromisoformat(args.fecha_inicio),
        dias_faena=args.dias_faena,
        pollos_por_dia=args.pollos_por_dia,
        criterio_gerente=True,
    )
    calendario = _resolver_calendario_planificacion(req, params, ofertas, [])
    semana = generar_proyeccion(
        ofertas=ofertas,
        fecha_inicio_semana=calendario["fecha_inicio"],
        dias_faena=calendario["dias_faena"],
        pollos_por_dia=req.pollos_por_dia,
        params=params,
        feriados=calendario["feriados"],
        criterio_gerente=True,
        permitir_fraccionamiento_lotes=True,
        excluir_backlog_semana_previa=True,
        minimos_como_alerta=True,
        planificacion_continua_gerente=calendario["planificacion_continua_gerente"],
    )
    return calendario, semana


def collect_projection_rows(semana) -> dict[LoteClave, list[dict]]:
    grouped: dict[LoteClave, list[dict]] = defaultdict(list)
    for dia in semana.dias:
        for lote in dia.lotes:
            clave = make_key(lote.granja, lote.galpon, lote.nucleo, lote.sexo)
            grouped[clave].append(
                {
                    "fecha": dia.fecha,
                    "cantidad": lote.cantidad,
                    "edad": lote.edad_fin_retiro,
                    "peso": lote.peso_vivo_retiro,
                }
            )
    for rows in grouped.values():
        rows.sort(key=lambda item: (item["fecha"], item["cantidad"]))
    return grouped


def first_date(rows: list[dict]) -> date | None:
    return rows[0]["fecha"] if rows else None


def summarize_keys(
    keys: Iterable[LoteClave],
    manager_rows: dict[LoteClave, list[dict]],
    projection_rows: dict[LoteClave, list[dict]],
) -> list[tuple[int, LoteClave]]:
    ranked = []
    for key in keys:
        manager_date = first_date(manager_rows.get(key, []))
        system_date = first_date(projection_rows.get(key, []))
        if manager_date and system_date:
            delta = abs((system_date - manager_date).days)
        elif manager_date or system_date:
            delta = 9999
        else:
            delta = -1
        ranked.append((delta, key))
    ranked.sort(key=lambda item: (-item[0], item[1].label()))
    return ranked


def fmt_rows(rows: list[dict]) -> str:
    if not rows:
        return "sin filas"
    parts = []
    for row in rows:
        parts.append(
            f"{row['fecha']} qty={row['cantidad']} edad={row['edad']} peso={row['peso']:.3f}"
        )
    return " | ".join(parts)


def main() -> None:
    args = parse_args()
    params = load_params(args.params, args.priorizar_peso_objetivo)
    calendario, semana = build_projection(args, params)
    manager_rows = load_manager_rows(args.manager, args.sheet)
    projection_rows = collect_projection_rows(semana)

    if args.keys:
        keys = [parse_key(raw) for raw in args.keys]
    else:
        universe = set(manager_rows) | set(projection_rows)
        keys = [key for _, key in summarize_keys(universe, manager_rows, projection_rows)[: args.top]]

    print("Calendario resuelto:")
    print(
        f"  fecha_inicio={calendario['fecha_inicio']} dias_faena={calendario['dias_faena']} "
        f"continuo={calendario['planificacion_continua_gerente']} "
        f"priorizar_peso_objetivo={params.planificacion_gerente_priorizar_peso_objetivo}"
    )
    print()

    for key in keys:
        manager = manager_rows.get(key, [])
        system = projection_rows.get(key, [])
        manager_date = first_date(manager)
        system_date = first_date(system)
        if manager_date and system_date:
            delta_text = f"delta_dias={(system_date - manager_date).days:+d}"
        else:
            delta_text = "delta_dias=n/a"
        print(key.label())
        print(f"  gerente: {fmt_rows(manager)}")
        print(f"  sistema: {fmt_rows(system)}")
        print(f"  {delta_text}")
        print()


if __name__ == "__main__":
    main()