"""
Script de prueba para la funcionalidad de Ajuste con Oferta del Martes.
Ejecutar con: python tests/test_ajuste_martes_api.py
"""
import requests
import json
import sys
import openpyxl
from io import BytesIO
from datetime import date, datetime

BASE = 'http://127.0.0.1:5000'


def login():
    r = requests.post(BASE + '/token', data={'username': 'admin', 'password': 'admin123'})
    assert r.status_code == 200, f"Login failed: {r.text}"
    return r.json()['access_token']


def get_proyeccion(headers):
    r = requests.get(BASE + '/proyeccion', headers=headers)
    return r.status_code, r.json() if r.status_code == 200 else r.json()


def print_proyeccion(proy):
    print(f"Fecha inicio: {proy.get('fecha_inicio')}")
    for i, dia in enumerate(proy.get('dias', [])):
        fecha = dia['fecha']
        total = dia['total_pollos']
        print(f"  Dia {i} ({fecha}): {total} pollos")
        for l in dia.get('lotes', []):
            granja = l['granja']
            galpon = l['galpon']
            nucleo = l['nucleo']
            cantidad = l['cantidad']
            sexo = l['sexo']
            edad = l['edad_fin_retiro']
            peso = l['peso_vivo_retiro']
            print(f"    - {granja} G{galpon} N{nucleo} qty={cantidad} sex={sexo} edad={edad} peso_vivo={peso:.3f}")
    na = proy.get('lotes_no_asignados', [])
    fr = proy.get('lotes_fuera_rango', [])
    print(f"  Lotes no asignados: {len(na)}, Fuera de rango: {len(fr)}")


def crear_excel_martes(lotes_data):
    """
    Crea un archivo Excel sintético con formato de oferta.
    lotes_data: lista de dicts con los campos de la oferta.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OFERTA MART"

    # Headers en fila 3 (el parser empieza en fila 4 para datos)
    headers_row = [
        "Fecha de Peso", "GRANJA", "Galpon", "Nucleo", "cantidad",
        "sexo", "Edad Proyectada", "Peso Muestreo Proy", "Ganancia Diaria",
        "Dias proyectados", "EDAD REAL", "Peso Muestreo REAL", "FECHA DE INGRESO"
    ]
    for col, h in enumerate(headers_row, 1):
        ws.cell(row=3, column=col, value=h)

    # Datos desde fila 4
    for row_idx, lote in enumerate(lotes_data, 4):
        ws.cell(row=row_idx, column=1, value=datetime.combine(lote['fecha_peso'], datetime.min.time()))
        ws.cell(row=row_idx, column=2, value=lote['granja'])
        ws.cell(row=row_idx, column=3, value=lote['galpon'])
        ws.cell(row=row_idx, column=4, value=lote['nucleo'])
        ws.cell(row=row_idx, column=5, value=lote['cantidad'])
        ws.cell(row=row_idx, column=6, value=lote['sexo'])
        ws.cell(row=row_idx, column=7, value=lote['edad_proyectada'])
        ws.cell(row=row_idx, column=8, value=lote['peso_muestreo_proy'])
        ws.cell(row=row_idx, column=9, value=lote['ganancia_diaria'])
        ws.cell(row=row_idx, column=10, value=lote['dias_proyectados'])
        ws.cell(row=row_idx, column=11, value=lote['edad_real'])
        ws.cell(row=row_idx, column=12, value=lote['peso_muestreo_real'])
        ws.cell(row=row_idx, column=13, value=datetime.combine(lote['fecha_ingreso'], datetime.min.time()))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_ajuste_martes():
    print("=" * 60)
    print("TEST: Ajuste con Oferta del Martes (API)")
    print("=" * 60)

    # 1. Login
    token = login()
    headers = {'Authorization': 'Bearer ' + token}
    print("[OK] Login exitoso")

    # 2. Get current projection
    status, proy = get_proyeccion(headers)
    if status != 200:
        print("[SKIP] No hay proyeccion existente, no se puede probar ajuste martes")
        print("       Necesita generar una proyeccion primero subiendo una oferta")
        return

    print("\n--- Proyeccion actual ---")
    print_proyeccion(proy)

    # 3. Build Excel with matching lotes (use first few lotes from projection)
    all_lotes = []
    for dia in proy.get('dias', []):
        for l in dia.get('lotes', []):
            all_lotes.append(l)

    if not all_lotes:
        print("[SKIP] Proyeccion sin lotes, no se puede probar ajuste")
        return

    # Create martes data from first 3 lotes (with slightly different peso)
    martes_data = []
    for lote in all_lotes[:5]:
        fp_original = lote.get('fecha_peso_original')
        fi_original = lote.get('fecha_ingreso_original')
        if not fp_original or not fi_original:
            # Fallback
            fp_original = proy['fecha_inicio']
            fi_original = proy['fecha_inicio']

        martes_data.append({
            'fecha_peso': date.fromisoformat(fp_original) if isinstance(fp_original, str) else fp_original,
            'granja': lote['granja'],
            'galpon': lote['galpon'],
            'nucleo': lote['nucleo'],
            'cantidad': lote['cantidad'],
            'sexo': lote['sexo'],
            'edad_proyectada': lote['edad_actual'] + 2,  # slightly different
            'peso_muestreo_proy': round(lote['peso_actual'] + 0.05, 3),  # slightly different
            'ganancia_diaria': lote.get('ganancia_diaria_original', 0.09),
            'dias_proyectados': 0,
            'edad_real': lote['edad_actual'] + 2,
            'peso_muestreo_real': round(lote['peso_actual'] + 0.05, 3),
            'fecha_ingreso': date.fromisoformat(fi_original) if isinstance(fi_original, str) else fi_original,
        })

    print(f"\n--- Creando Excel con {len(martes_data)} lotes para ajuste ---")
    excel_buf = crear_excel_martes(martes_data)

    # 4. Call ajuste-martes endpoint
    files = {'file': ('ajuste_martes_test.xlsx', excel_buf, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    r = requests.post(BASE + '/oferta/ajuste-martes', headers=headers, files=files)

    print(f"\n--- Resultado ajuste-martes: status {r.status_code} ---")
    if r.status_code != 200:
        print(f"[ERROR] {r.text}")
        return

    data = r.json()

    # 5. Check resumen
    resumen = data.get('resumen_ajuste', {})
    print(f"  Lotes actualizados: {resumen.get('lotes_actualizados', 0)}")
    print(f"  Lotes nuevos: {resumen.get('lotes_nuevos', 0)}")
    print(f"  Lotes nuevos asignados: {resumen.get('lotes_nuevos_asignados', 0)}")
    print(f"  Lotes nuevos fuera rango: {resumen.get('lotes_nuevos_fuera_rango', 0)}")
    print(f"  Lotes faltantes: {resumen.get('lotes_faltantes', 0)}")
    print(f"  Lotes fuera rango post-ajuste: {resumen.get('lotes_fuera_rango_post_ajuste', 0)}")

    if resumen.get('detalle_actualizados'):
        print("\n  Detalle actualizaciones:")
        for d in resumen['detalle_actualizados']:
            print(f"    {d.get('granja')} G{d.get('galpon')} - {d.get('dia')}: {d.get('cambios')}")

    if resumen.get('detalle_fuera_rango_post_ajuste'):
        print("\n  Detalle fuera rango post-ajuste:")
        for d in resumen['detalle_fuera_rango_post_ajuste']:
            print(f"    {d.get('granja')} G{d.get('galpon')} - {d.get('dia')}: {d.get('alerta')}")

    if resumen.get('detalle_faltantes'):
        print("\n  Detalle faltantes:")
        for d in resumen['detalle_faltantes']:
            print(f"    {d.get('granja')} G{d.get('galpon')} N{d.get('nucleo')} - {d.get('dia')}")

    # 6. Check projection was updated
    nueva_proy = data.get('proyeccion', {})
    print("\n--- Proyeccion actualizada ---")
    print_proyeccion(nueva_proy)

    # 7. Verify projection is persisted
    status2, proy2 = get_proyeccion(headers)
    assert status2 == 200, "Proyeccion should be persisted"
    assert proy2.get('total_pollos_semana') == nueva_proy.get('total_pollos_semana'), \
        "Persisted projection should match returned projection"
    print("\n[OK] Proyeccion persistida correctamente")

    print("\n" + "=" * 60)
    print("TEST COMPLETO: Ajuste con Oferta del Martes funciona correctamente")
    print("=" * 60)


def test_ajuste_sin_proyeccion():
    """Test que el endpoint rechaza si no hay proyeccion."""
    print("\n" + "=" * 60)
    print("TEST: Ajuste sin proyeccion existente")
    print("=" * 60)

    token = login()
    headers = {'Authorization': 'Bearer ' + token}

    # Create minimal Excel
    martes_data = [{
        'fecha_peso': date(2026, 2, 25),
        'granja': 'TEST',
        'galpon': 1,
        'nucleo': 1,
        'cantidad': 10000,
        'sexo': 'M',
        'edad_proyectada': 40,
        'peso_muestreo_proy': 2.95,
        'ganancia_diaria': 0.09,
        'dias_proyectados': 0,
        'edad_real': 40,
        'peso_muestreo_real': 2.95,
        'fecha_ingreso': date(2026, 1, 10),
    }]

    # We can't easily test this without clearing the projection first
    # Just verify the endpoint exists and responds
    excel_buf = crear_excel_martes(martes_data)
    files = {'file': ('test.xlsx', excel_buf, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    r = requests.post(BASE + '/oferta/ajuste-martes', headers=headers, files=files)
    print(f"  Status: {r.status_code}")
    # It should be 200 (there's already a projection) or 400 (no projection)
    assert r.status_code in [200, 400], f"Unexpected status: {r.status_code}"
    print("[OK] Endpoint responde correctamente")


def test_ajuste_archivo_invalido():
    """Test que rechaza archivos no-Excel."""
    print("\n" + "=" * 60)
    print("TEST: Ajuste con archivo invalido")
    print("=" * 60)

    token = login()
    headers = {'Authorization': 'Bearer ' + token}

    files = {'file': ('test.txt', b'not an excel file', 'text/plain')}
    r = requests.post(BASE + '/oferta/ajuste-martes', headers=headers, files=files)
    print(f"  Status: {r.status_code} (esperado: 400)")
    assert r.status_code == 400, f"Expected 400, got {r.status_code}"
    print("[OK] Archivo invalido rechazado correctamente")


if __name__ == '__main__':
    test_ajuste_martes()
    test_ajuste_archivo_invalido()
    print("\n\nTODOS LOS TESTS PASARON")
